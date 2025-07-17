# -*- coding: utf-8 -*-
"""
Excel 输出器
将筛选后的推文数据导出到 Excel 文件
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_CONFIG

class ExcelWriter:
    def __init__(self):
        self.data_dir = OUTPUT_CONFIG['data_dir']
        self.filename_format = OUTPUT_CONFIG['excel_filename_format']
        self.sheet_name = OUTPUT_CONFIG['sheet_name']
        self.logger = logging.getLogger(__name__)
        
        # 确保数据目录存在
        os.makedirs(self.data_dir, exist_ok=True)
    
    def generate_filename(self, date: Optional[str] = None) -> str:
        """
        生成 Excel 文件名
        
        Args:
            date: 日期字符串，如果不提供则使用当前日期
            
        Returns:
            完整的文件路径
        """
        if not date:
            date = datetime.now().strftime('%Y%m%d')
        
        filename = self.filename_format.format(date=date)
        return os.path.join(self.data_dir, filename)
    
    def create_workbook(self) -> Workbook:
        """
        创建新的工作簿
        
        Returns:
            Excel 工作簿对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        return wb
    
    def setup_header_style(self, ws, header_row: int = 1):
        """
        设置表头样式
        
        Args:
            ws: 工作表对象
            header_row: 表头行号
        """
        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用样式到表头行
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
    def setup_data_style(self, ws, start_row: int = 2):
        """
        设置数据行样式
        
        Args:
            ws: 工作表对象
            start_row: 数据开始行号
        """
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用样式到数据行
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                # 交替行颜色
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    def auto_adjust_column_width(self, ws):
        """
        自动调整列宽
        
        Args:
            ws: 工作表对象
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置最小和最大宽度
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def format_tweet_content(self, content: str, max_length: int = 100) -> str:
        """
        格式化推文内容
        
        Args:
            content: 原始推文内容
            max_length: 最大长度
            
        Returns:
            格式化后的内容
        """
        if not content:
            return ''
        
        # 移除多余的空白字符
        content = ' '.join(content.split())
        
        # 截断过长的内容
        if len(content) > max_length:
            content = content[:max_length] + '...'
        
        return content
    
    def format_datetime(self, datetime_str: str) -> str:
        """
        格式化日期时间
        
        Args:
            datetime_str: ISO 格式的日期时间字符串
            
        Returns:
            格式化后的日期时间字符串
        """
        try:
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return datetime_str
    
    def write_tweets_to_excel(self, tweets: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """
        将推文数据写入 Excel 文件
        
        Args:
            tweets: 推文数据列表
            filename: 输出文件名，如果不提供则自动生成
            
        Returns:
            生成的文件路径
        """
        if not tweets:
            self.logger.warning("没有推文数据需要写入 Excel")
            return ''
        
        # 生成文件名
        if not filename:
            filename = self.generate_filename()
        
        self.logger.info(f"开始写入 {len(tweets)} 条推文数据到 Excel: {filename}")
        
        try:
            # 创建工作簿
            wb = self.create_workbook()
            ws = wb.active
            
            # 定义列标题
            headers = [
                '序号',
                '账号',
                '推文内容',
                '发布时间',
                '评论数',
                '转发数',
                '点赞数',
                '推文链接',
                '来源',
                '来源类型',
                '筛选原因'
            ]
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, tweet in enumerate(tweets, 2):
                ws.cell(row=row, column=1, value=row - 1)  # 序号
                ws.cell(row=row, column=2, value=f"@{tweet.get('username', '')}")
                ws.cell(row=row, column=3, value=self.format_tweet_content(tweet.get('content', '')))
                ws.cell(row=row, column=4, value=self.format_datetime(tweet.get('publish_time', '')))
                ws.cell(row=row, column=5, value=tweet.get('comments', 0))
                ws.cell(row=row, column=6, value=tweet.get('retweets', 0))
                ws.cell(row=row, column=7, value=tweet.get('likes', 0))
                ws.cell(row=row, column=8, value=tweet.get('link', ''))
                ws.cell(row=row, column=9, value=tweet.get('source', ''))
                ws.cell(row=row, column=10, value=tweet.get('source_type', ''))
                ws.cell(row=row, column=11, value=', '.join(tweet.get('filter_reasons', [])))
            
            # 应用样式
            self.setup_header_style(ws)
            self.setup_data_style(ws)
            self.auto_adjust_column_width(ws)
            
            # 保存文件
            wb.save(filename)
            self.logger.info(f"Excel 文件保存成功: {filename}")
            
            return filename
            
        except Exception as e:
            self.logger.error(f"写入 Excel 文件失败: {e}")
            raise
    
    def create_summary_sheet(self, wb: Workbook, statistics: Dict[str, Any]):
        """
        创建汇总统计表
        
        Args:
            wb: 工作簿对象
            statistics: 统计信息
        """
        try:
            # 创建汇总表
            summary_ws = wb.create_sheet(title='汇总统计')
            
            # 基本统计信息
            summary_ws.cell(row=1, column=1, value='统计项目')
            summary_ws.cell(row=1, column=2, value='数值')
            
            row = 2
            summary_ws.cell(row=row, column=1, value='总推文数')
            summary_ws.cell(row=row, column=2, value=statistics.get('total_tweets', 0))
            
            row += 1
            summary_ws.cell(row=row, column=1, value='通过筛选推文数')
            summary_ws.cell(row=row, column=2, value=statistics.get('passed_tweets', 0))
            
            row += 1
            summary_ws.cell(row=row, column=1, value='筛选通过率')
            summary_ws.cell(row=row, column=2, value=f"{statistics.get('pass_rate', 0):.2%}")
            
            # 互动数据统计
            engagement_stats = statistics.get('engagement_stats', {})
            row += 2
            summary_ws.cell(row=row, column=1, value='总点赞数')
            summary_ws.cell(row=row, column=2, value=engagement_stats.get('total_likes', 0))
            
            row += 1
            summary_ws.cell(row=row, column=1, value='总评论数')
            summary_ws.cell(row=row, column=2, value=engagement_stats.get('total_comments', 0))
            
            row += 1
            summary_ws.cell(row=row, column=1, value='总转发数')
            summary_ws.cell(row=row, column=2, value=engagement_stats.get('total_retweets', 0))
            
            row += 1
            summary_ws.cell(row=row, column=1, value='平均点赞数')
            summary_ws.cell(row=row, column=2, value=f"{engagement_stats.get('avg_likes', 0):.1f}")
            
            row += 1
            summary_ws.cell(row=row, column=1, value='平均评论数')
            summary_ws.cell(row=row, column=2, value=f"{engagement_stats.get('avg_comments', 0):.1f}")
            
            row += 1
            summary_ws.cell(row=row, column=1, value='平均转发数')
            summary_ws.cell(row=row, column=2, value=f"{engagement_stats.get('avg_retweets', 0):.1f}")
            
            # 筛选原因统计
            filter_reasons = statistics.get('filter_reasons', {})
            if filter_reasons:
                row += 2
                summary_ws.cell(row=row, column=1, value='筛选原因统计')
                row += 1
                for reason, count in filter_reasons.items():
                    summary_ws.cell(row=row, column=1, value=reason)
                    summary_ws.cell(row=row, column=2, value=count)
                    row += 1
            
            # 应用样式
            self.setup_header_style(summary_ws, 1)
            self.auto_adjust_column_width(summary_ws)
            
            self.logger.info("汇总统计表创建成功")
            
        except Exception as e:
            self.logger.error(f"创建汇总统计表失败: {e}")
    
    def write_tweets_with_summary(self, tweets: List[Dict[str, Any]], statistics: Dict[str, Any], 
                                 filename: Optional[str] = None) -> str:
        """
        将推文数据和统计信息写入 Excel 文件
        
        Args:
            tweets: 推文数据列表
            statistics: 统计信息
            filename: 输出文件名
            
        Returns:
            生成的文件路径
        """
        if not tweets:
            self.logger.warning("没有推文数据需要写入 Excel")
            return ''
        
        # 生成文件名
        if not filename:
            filename = self.generate_filename()
        
        self.logger.info(f"开始写入 {len(tweets)} 条推文数据和统计信息到 Excel: {filename}")
        
        try:
            # 创建工作簿并写入推文数据
            wb = self.create_workbook()
            ws = wb.active
            
            # 定义列标题
            headers = [
                '序号', '账号', '推文内容', '发布时间', '点赞数', '评论数', '转发数',
                '推文链接', '来源', '来源类型', '筛选原因'
            ]
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, tweet in enumerate(tweets, 2):
                ws.cell(row=row, column=1, value=row - 1)
                ws.cell(row=row, column=2, value=f"@{tweet.get('username', '')}")
                ws.cell(row=row, column=3, value=self.format_tweet_content(tweet.get('content', '')))
                ws.cell(row=row, column=4, value=self.format_datetime(tweet.get('publish_time', '')))
                ws.cell(row=row, column=5, value=tweet.get('likes', 0))
                ws.cell(row=row, column=6, value=tweet.get('comments', 0))
                ws.cell(row=row, column=7, value=tweet.get('retweets', 0))
                ws.cell(row=row, column=8, value=tweet.get('link', ''))
                ws.cell(row=row, column=9, value=tweet.get('source', ''))
                ws.cell(row=row, column=10, value=tweet.get('source_type', ''))
                ws.cell(row=row, column=11, value=', '.join(tweet.get('filter_reasons', [])))
            
            # 应用样式
            self.setup_header_style(ws)
            self.setup_data_style(ws)
            self.auto_adjust_column_width(ws)
            
            # 创建汇总统计表
            self.create_summary_sheet(wb, statistics)
            
            # 保存文件
            wb.save(filename)
            self.logger.info(f"Excel 文件（含统计信息）保存成功: {filename}")
            
            return filename
            
        except Exception as e:
            self.logger.error(f"写入 Excel 文件失败: {e}")
            raise

# 使用示例
if __name__ == "__main__":
    # 配置日志
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # 示例推文数据
    sample_tweets = [
        {
            'username': 'elonmusk',
            'content': 'AI is the future of humanity and will change everything we know about technology',
            'likes': 1500,
            'comments': 200,
            'retweets': 800,
            'publish_time': '2024-01-01T12:00:00Z',
            'link': 'https://twitter.com/elonmusk/status/123',
            'source': '@elonmusk',
            'source_type': 'user_profile',
            'filter_passed': True,
            'filter_reasons': ['满足互动数阈值', '包含目标关键词']
        },
        {
            'username': 'techuser',
            'content': '今天学习了一个新的副业技巧，分享给大家',
            'likes': 80,
            'comments': 35,
            'retweets': 15,
            'publish_time': '2024-01-01T14:00:00Z',
            'link': 'https://twitter.com/techuser/status/125',
            'source': '副业',
            'source_type': 'keyword_search',
            'filter_passed': True,
            'filter_reasons': ['满足互动数阈值', '包含目标关键词']
        }
    ]
    
    # 示例统计信息
    sample_statistics = {
        'total_tweets': 10,
        'passed_tweets': 2,
        'pass_rate': 0.2,
        'filter_reasons': {
            '满足互动数阈值': 2,
            '包含目标关键词': 2
        },
        'engagement_stats': {
            'total_likes': 1580,
            'total_comments': 235,
            'total_retweets': 815,
            'avg_likes': 158.0,
            'avg_comments': 23.5,
            'avg_retweets': 81.5
        }
    }
    
    # 创建 Excel 输出器
    excel_writer = ExcelWriter()
    
    try:
        # 写入 Excel 文件
        output_file = excel_writer.write_tweets_with_summary(sample_tweets, sample_statistics)
        print(f"Excel 文件已生成: {output_file}")
        
    except Exception as e:
        print(f"错误: {e}")